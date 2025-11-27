#!/usr/bin/env python3
"""
Text File CHF to EUR Converter
Converts CHF prices to EUR in text files using Stock Lines.xlsx matching
Also generates filtered Stock Lines Excel with recognized items
"""

import re
import pandas as pd
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import math
import requests
import openpyxl
from shutil import copy2

# Configuration
BASE_DIR = r"C:\Users\Marco.Africani\Desktop\Month recap"
DATABASE_DIR = r"C:\Users\Marco.Africani\OneDrive - AVU SA\AVU CPI Campaign\Puzzle_control_Reports\SOURCE_FILES"
INPUT_FILE_PATH = rf"{BASE_DIR}\Inputs\Multi.txt"
EXCEL_FILE_PATH = rf"{DATABASE_DIR}\OMT Main Offer List.xlsx"
STOCK_FILE_PATH = rf"{DATABASE_DIR}\Stock Lines.xlsx"
OUTPUTS_DIR = rf"{BASE_DIR}\Outputs"
RECOGNITION_REPORT_DIR = rf"{BASE_DIR}\Outputs\Detailed match results"
TRANSLATIONS_DIR = rf"{BASE_DIR}\Outputs\EUR translations"
MAIN_OFFER_DIR = rf"{BASE_DIR}\Outputs\Detailed match results\Main offer"
LINES_TEMPLATE_PATH = rf"{MAIN_OFFER_DIR}\template\Lines Template.xlsx"

# DeepL API Configuration
DEEPL_API_KEY = "374a8965-101a-4538-bc65-54506552650e"
DEEPL_API_URL = "https://api.deepl.com/v2/translate"

# Regex patterns for price detection
# Pattern for CHF prices - handles multiple formats:
# "CHF 900.00", "230.00 CHF", "29.00 CH + VAT", "CHF 50.00"
CHF_PRICE_PATTERN = r'(?:\bCHF\s+(\d+(?:\.\d{2})?)|(\d+(?:\.\d{2})?)\s+CHF\b|(\d+(?:\.\d{2})?)\s+CH\s)'

# Pattern for EUR prices (for output)
EUR_PRICE_PATTERN = r'\bEUR\s+(\d+(?:\.\d{2})?)\b'


def load_stock_lines():
    """Load Stock Lines.xlsx for direct Item No. to EUR price matching"""
    try:
        print(f"Loading Stock Lines database from: {STOCK_FILE_PATH}")
        stock_df = pd.read_excel(STOCK_FILE_PATH)
        print(f"[OK] Loaded {len(stock_df)} items from Stock Lines.xlsx")
        return stock_df
    except FileNotFoundError:
        print(f"[WARN] Stock Lines.xlsx not found at {STOCK_FILE_PATH}")
        return None
    except Exception as e:
        print(f"[WARN] Error loading Stock Lines.xlsx: {e}")
        return None


def load_omt_data():
    """Load OMT Main Offer List for EUR price lookups"""
    try:
        print(f"Loading OMT Main Offer List from: {EXCEL_FILE_PATH}")
        df = pd.read_excel(EXCEL_FILE_PATH)
        print(f"[OK] Loaded {len(df)} rows from OMT Main Offer List.xlsx")
        return df
    except FileNotFoundError:
        print(f"[WARN] OMT Main Offer List.xlsx not found at {EXCEL_FILE_PATH}")
        return None
    except Exception as e:
        print(f"[WARN] Error loading OMT Main Offer List.xlsx: {e}")
        return None


def extract_wine_name_and_vintage_from_line(line_text):
    """
    Extract wine name and vintage from a single line

    Expected formats:
    - "Château Wine Name VINTAGE - Region: description CHF price"
    - "Wine Name VINTAGE: description CHF price"
    - "Wine Name VINTAGE – description – at CHF price"
    - "Wine Name VINTAGE at CHF price"

    The wine name and vintage are at the START of each line, before the first CHF price.

    Args:
        line_text: Single line of text (one wine entry)

    Returns:
        tuple: (wine_name, vintage)
    """
    # Get text before the first CHF price (this should contain the wine name and vintage)
    chf_pos = line_text.find('CHF')
    if chf_pos > 0:
        wine_part = line_text[:chf_pos].strip()
    else:
        wine_part = line_text[:150].strip()  # Fallback: first 150 chars

    # Pattern 1: "Wine Name VINTAGE:" or "Wine Name VINTAGE -" or "Wine Name VINTAGE –"
    # This matches: "Château Pavie 2019 -", "Oreno 2023:", "Figeac 2018:", "Cos d'Estournel 2021 –"
    # Allow hyphens, apostrophes, and ANY unicode character in wine name
    # This handles encoding issues with special characters
    pattern1 = r'^\s*([^\d]+?)\s+(19\d{2}|20\d{2})\s*[^\w\s]'
    match = re.search(pattern1, wine_part)

    if match:
        wine_name_raw = match.group(1).strip()
        vintage = int(match.group(2))

        # Clean up wine name - remove any trailing digits and spaces
        wine_name = re.sub(r'\s*\d+\s*$', '', wine_name_raw)
        wine_name = re.sub(r'^[-–—\s]+', '', wine_name)
        wine_name = re.sub(r'^(Château|Chateau|Domaine|Dom\.|Ch\.)\s+', '', wine_name, flags=re.IGNORECASE)

        return wine_name, vintage

    # Pattern 2: "Wine Name VINTAGE at CHF" or "Wine Name VINTAGE instead"
    # This matches: "Pavie Maquin 2021 at CHF", "Mondotte 2021 – one"
    pattern2 = r'^\s*([A-Za-zÀ-ÿ\s\.\'\-]+?)\s+(19\d{2}|20\d{2})\s+(?:at|instead|–)'
    match = re.search(pattern2, wine_part)

    if match:
        wine_name = match.group(1).strip()
        vintage = int(match.group(2))

        # Clean up wine name
        wine_name = re.sub(r'^[-–—\s]+', '', wine_name)
        wine_name = re.sub(r'^(Château|Chateau|Domaine|Dom\.|Ch\.)\s+', '', wine_name, flags=re.IGNORECASE)

        return wine_name, vintage

    # Pattern 3: Just "Wine Name VINTAGE" at start with whitespace after
    pattern3 = r'^\s*([A-Za-zÀ-ÿ\s\.\'\-]+?)\s+(19\d{2}|20\d{2})\s'
    match = re.search(pattern3, wine_part)

    if match:
        wine_name = match.group(1).strip()
        vintage = int(match.group(2))

        # Clean up wine name
        wine_name = re.sub(r'^[-–—\s]+', '', wine_name)
        wine_name = re.sub(r'^(Château|Chateau|Domaine|Dom\.|Ch\.)\s+', '', wine_name, flags=re.IGNORECASE)

        return wine_name, vintage

    return None, None


def translate_text_deepl(text, target_language):
    """
    Translate text using DeepL API

    Args:
        text: Text to translate
        target_language: Target language code ('DE' for German, 'FR' for French)

    Returns:
        Translated text or None if error
    """
    try:
        response = requests.post(
            DEEPL_API_URL,
            data={
                'auth_key': DEEPL_API_KEY,
                'text': text,
                'target_lang': target_language
            },
            timeout=30
        )

        if response.status_code == 200:
            result = response.json()
            return result['translations'][0]['text']
        else:
            print(f"[WARN] DeepL API error: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"[WARN] Translation error for {target_language}: {e}")
        return None


def match_price_via_stock_lines(chf_price, wine_name, vintage, stock_df, omt_df, min_quantity=0):
    """
    BULLETPROOF MATCHING LOGIC:

    Stock Lines "OMT Last Private Offer Price" = OMT "Unit Price" (CHF price used by author)

    Steps:
    1. Find Stock Lines items where "OMT Last Private Offer Price" matches CHF price
    2. Get Item No. from Stock Lines
    3. Match in OMT by: Item No. + Unit Price (CHF) + Minimum Quantity
    4. If multiple matches, pick latest Schedule DateTime
    5. Return EUR price from OMT "Unit Price (EUR)"

    Example: Oreno 2023
    - Stock Lines row 3381: Item 65245, OMT Last Private Offer Price = 49.00
    - OMT row 3094: Item 65245, Unit Price = 49.00, Minimum Quantity = 0
    - EUR price from OMT: Unit Price (EUR) column J

    Args:
        chf_price: CHF price as float
        wine_name: Wine name extracted from text (for disambiguation if needed)
        vintage: Vintage year (for disambiguation if needed)
        stock_df: Stock Lines DataFrame
        omt_df: OMT Main Offer List DataFrame
        min_quantity: Minimum quantity filter (0 for standard, 36 for bulk)

    Returns:
        dict with item info including EUR price or None
    """
    if stock_df is None or omt_df is None:
        return None

    try:
        chf_float = float(chf_price)

        # Step 1: Find in Stock Lines by OMT Last Private Offer Price (CHF)
        stock_matches = stock_df[
            stock_df['OMT Last Private Offer Price'].astype(float).round(2) == round(chf_float, 2)
        ].copy()

        if len(stock_matches) == 0:
            return None

        # Step 2: For each Stock Lines match, find in OMT by Item No. + Unit Price + Min Quantity
        all_candidates = []

        for idx, stock_row in stock_matches.iterrows():
            item_no = stock_row['No.']

            # Step 3: Find in OMT Main Offer List by Item No., Unit Price (CHF), and Minimum Quantity
            omt_df_copy = omt_df.copy()
            omt_df_copy['Item No. Int'] = pd.to_numeric(omt_df_copy['Item No.'], errors='coerce').astype('Int64')
            omt_df_copy['Schedule DateTime'] = pd.to_datetime(omt_df_copy['Schedule DateTime'])

            # BULLETPROOF MATCH: Item No. + Unit Price (CHF) + Minimum Quantity
            omt_matches = omt_df_copy[
                (omt_df_copy['Item No. Int'] == item_no) &
                (omt_df_copy['Unit Price'].astype(float).round(2) == round(chf_float, 2)) &
                (omt_df_copy['Minimum Quantity'].astype(float) == float(min_quantity))
            ]

            # Add all matches to candidates
            for _, omt_row in omt_matches.iterrows():
                all_candidates.append({
                    'item_no': item_no,
                    'chf_price': chf_price,
                    'eur_price': omt_row.get('Unit Price (EUR)', 0),
                    'stock_row': stock_row,
                    'omt_row': omt_row,
                    'schedule_datetime': omt_row['Schedule DateTime']
                })

        if len(all_candidates) == 0:
            return None

        # Step 4: Pick latest Schedule DateTime if multiple matches
        if len(all_candidates) > 1:
            all_candidates.sort(key=lambda x: x['schedule_datetime'], reverse=True)

        # Step 5: Return best match (latest datetime)
        best_match = all_candidates[0]

        # Optional: Score by wine name/vintage for additional disambiguation
        if wine_name and len(all_candidates) > 1:
            for candidate in all_candidates:
                omt_wine = str(candidate['omt_row'].get('Wine Name', '')).lower().strip()
                omt_vintage = candidate['omt_row'].get('Vintage', None)
                score = 0

                # Wine name match
                wine_name_lower = wine_name.lower().strip()
                if wine_name_lower in omt_wine or omt_wine in wine_name_lower:
                    score += 20

                # Vintage match
                if vintage and omt_vintage:
                    try:
                        if int(omt_vintage) == vintage:
                            score += 20
                    except:
                        pass

                candidate['name_vintage_score'] = score

            # Re-sort by name/vintage score if scores differ
            all_candidates.sort(key=lambda x: (x.get('name_vintage_score', 0), x['schedule_datetime']), reverse=True)
            best_match = all_candidates[0]

        # Remove temporary fields
        best_match.pop('schedule_datetime', None)
        best_match.pop('name_vintage_score', None)

        return best_match

    except Exception as e:
        print(f"[WARN] Error matching price {chf_price}: {e}")
        import traceback
        traceback.print_exc()
        return None


def match_eur_price_via_omt(eur_price, wine_name, vintage, stock_df, omt_df, min_quantity=0):
    """
    Match EUR price directly in OMT Main Offer List, then get Stock Lines data

    For files that already have EUR prices instead of CHF

    Args:
        eur_price: EUR price as float
        wine_name: Wine name extracted from text
        vintage: Vintage year
        stock_df: Stock Lines DataFrame
        omt_df: OMT Main Offer List DataFrame
        min_quantity: Minimum quantity filter (0 for standard, 36 for bulk)

    Returns:
        dict with item info including CHF and EUR prices or None
    """
    if stock_df is None or omt_df is None:
        return None

    try:
        eur_float = float(eur_price)

        # Step 1: Find in OMT by EUR price, Size, and Minimum Quantity
        omt_df_copy = omt_df.copy()
        omt_df_copy['Item No. Int'] = pd.to_numeric(omt_df_copy['Item No.'], errors='coerce').astype('Int64')

        omt_matches = omt_df_copy[
            (omt_df_copy['Unit Price (EUR)'].astype(float).round(2) == round(eur_float, 2)) &
            (omt_df_copy['Size'].astype(float) == 75.0) &
            (omt_df_copy['Minimum Quantity'].astype(float) == float(min_quantity))
        ]

        if len(omt_matches) == 0:
            return None

        # Step 2: Score matches by wine name and vintage
        best_match = None
        best_score = 0

        for _, omt_row in omt_matches.iterrows():
            score = 100  # Start with base score

            omt_wine = str(omt_row.get('Wine Name', '')).lower().strip()
            omt_vintage = omt_row.get('Vintage', None)

            # Wine name similarity bonus
            if wine_name:
                wine_name_lower = wine_name.lower().strip()
                if wine_name_lower in omt_wine or omt_wine in wine_name_lower:
                    score += 20
                # Partial word match
                wine_words = wine_name_lower.split()
                for word in wine_words:
                    if len(word) > 3 and word in omt_wine:
                        score += 5

            # Vintage match bonus
            if vintage and omt_vintage:
                try:
                    if int(omt_vintage) == vintage:
                        score += 20
                except:
                    pass

            if score > best_score:
                best_score = score
                item_no = int(omt_row['Item No. Int'])

                # Step 3: Find in Stock Lines
                stock_match = stock_df[stock_df['No.'] == item_no]

                if len(stock_match) > 0:
                    best_match = {
                        'item_no': item_no,
                        'chf_price': omt_row.get('Unit Price', 0),
                        'eur_price': eur_price,
                        'stock_row': stock_match.iloc[0],
                        'omt_row': omt_row
                    }

        return best_match

    except Exception as e:
        print(f"[WARN] Error matching EUR price {eur_price}: {e}")
        import traceback
        traceback.print_exc()
        return None


def convert_txt_file(enable_translations=True):
    """Main conversion function for txt files"""
    print("\n" + "="*80)
    print("TXT File CHF to EUR Converter")
    print("="*80 + "\n")

    # Load databases
    stock_df = load_stock_lines()
    omt_df = load_omt_data()

    if stock_df is None or omt_df is None:
        print("\n[ERROR] Cannot proceed without database files")
        return

    # Read input text file
    try:
        with open(INPUT_FILE_PATH, 'r', encoding='utf-8') as f:
            text = f.read()
        print(f"[OK] Loaded input file: {INPUT_FILE_PATH}")
    except FileNotFoundError:
        print(f"[ERROR] Input file not found: {INPUT_FILE_PATH}")
        return
    except Exception as e:
        print(f"[ERROR] Error reading input file: {e}")
        return

    # NEW APPROACH: Split text into paragraphs and process each paragraph
    # Each paragraph typically contains: Wine Name VINTAGE: description CHF price(s)
    recognized_items = []
    all_matches = []
    converted_text = text  # Start with original text

    # Check if input has CHF or EUR prices
    chf_matches = list(re.finditer(CHF_PRICE_PATTERN, text))
    eur_matches = list(re.finditer(EUR_PRICE_PATTERN, text))

    if len(chf_matches) > 0:
        # Input has CHF prices - convert to EUR
        price_pattern = CHF_PRICE_PATTERN
        price_currency = "CHF"
    elif len(eur_matches) > 0:
        # Input already has EUR prices - match directly
        price_pattern = EUR_PRICE_PATTERN
        price_currency = "EUR"
    else:
        print("[WARN] No CHF or EUR prices found in input file")
        return

    print(f"[INFO] Detected {price_currency} prices in input file")

    # Split text into lines (each wine is on its own line)
    lines = text.split('\n')

    # Collect all CHF->EUR replacements first, then apply in reverse order
    replacements = []  # List of (start_pos, end_pos, replacement_text)

    for match in re.finditer(price_pattern, text):
        # Extract price from whichever group matched (1, 2, or 3)
        price_str = match.group(1) or match.group(2) or match.group(3)
        if not price_str:
            continue
        price_value = float(price_str)
        position = match.start()

        # Find which line contains this price
        line_text = None
        cumulative_pos = 0
        for line in lines:
            line_start = cumulative_pos
            line_end = cumulative_pos + len(line)
            if line_start <= position <= line_end:
                line_text = line
                break
            cumulative_pos = line_end + 1  # Account for \n separator

        # If no line found, use context window
        if line_text is None:
            line_text = text[max(0, position-300):min(len(text), position+50)]

        # Extract wine name and vintage from line (before the CHF price)
        wine_name, vintage = extract_wine_name_and_vintage_from_line(line_text)

        # For EUR prices, we need to find CHF price by reverse lookup
        if price_currency == "EUR":
            # Match EUR price directly in OMT, then get CHF price
            item_info = match_eur_price_via_omt(price_value, wine_name, vintage, stock_df, omt_df, min_quantity=0)
            if item_info is None:
                item_info = match_eur_price_via_omt(price_value, wine_name, vintage, stock_df, omt_df, min_quantity=36)
        else:
            # CHF to EUR conversion (original logic)
            item_info = match_price_via_stock_lines(price_value, wine_name, vintage, stock_df, omt_df, min_quantity=0)
            if item_info is None:
                item_info = match_price_via_stock_lines(price_value, wine_name, vintage, stock_df, omt_df, min_quantity=36)

        if item_info:
            # Store wine name and vintage from Multi.txt, not from OMT database
            item_info['extracted_wine_name'] = wine_name
            item_info['extracted_vintage'] = vintage
            recognized_items.append(item_info)
            chf_price = item_info['chf_price']
            eur_price = item_info['eur_price']
            # Remove non-ASCII characters for printing
            safe_wine_name = wine_name.encode('ascii', errors='ignore').decode('ascii') if wine_name else ''

            if price_currency == "EUR":
                print(f"[OK] Matched: {safe_wine_name} {vintage} - EUR {eur_price} (CHF {chf_price}) - Item No. {item_info['item_no']}")
                # EUR prices don't need conversion - text stays as-is
            else:
                print(f"[OK] Matched: {safe_wine_name} {vintage} - CHF {chf_price} -> EUR {eur_price} (Item No. {item_info['item_no']})")

                # Prepare EUR replacement text
                eur_rounded = f"{int(round(float(eur_price)))}.00"
                original_match_text = match.group(0)

                # Create EUR replacement maintaining the same format
                if 'CHF' in original_match_text:
                    eur_replacement = original_match_text.replace('CHF', 'EUR')
                    eur_replacement = re.sub(r'\d+\.\d{2}', eur_rounded, eur_replacement)
                elif 'CH ' in original_match_text:
                    eur_replacement = original_match_text.replace('CH ', 'EUR ')
                    eur_replacement = re.sub(r'\d+\.\d{2}', eur_rounded, eur_replacement)
                else:
                    # Price without CHF keyword, just replace the number
                    eur_replacement = original_match_text.replace(price_str, eur_rounded)

                # Store replacement to apply later (in reverse order)
                replacements.append((match.start(), match.end(), eur_replacement))

            all_matches.append({
                'wine_name': wine_name,
                'vintage': vintage,
                'chf_price': chf_price,
                'eur_price': eur_price,
                'matched': True
            })
        else:
            safe_wine_name = wine_name.encode('ascii', errors='ignore').decode('ascii') if wine_name else ''

            if price_currency == "EUR":
                print(f"[WARN] Not matched: {safe_wine_name} {vintage} - EUR {price_value}")
                # Can't do anything without match for EUR
                all_matches.append({
                    'wine_name': wine_name,
                    'vintage': vintage,
                    'chf_price': None,
                    'eur_price': price_value,
                    'matched': False
                })
            else:
                # Fallback: use 1.08 conversion for CHF
                eur_fallback = price_value * 1.08
                print(f"[WARN] Not matched: {safe_wine_name} {vintage} - CHF {price_value} (using 1.08 fallback)")

                # Replace CHF with EUR using fallback
                price_pattern_text = f"CHF {price_str}"
                eur_replacement = f"EUR {int(math.floor(eur_fallback))}.00"
                converted_text = converted_text.replace(price_pattern_text, eur_replacement, 1)

                all_matches.append({
                    'wine_name': wine_name,
                    'vintage': vintage,
                    'chf_price': price_value,
                    'eur_price': eur_fallback,
                    'matched': False
                })

    # Apply all CHF->EUR replacements in reverse order (to maintain positions)
    if price_currency == "CHF":
        replacements.sort(key=lambda x: x[0], reverse=True)
        for start_pos, end_pos, replacement_text in replacements:
            converted_text = converted_text[:start_pos] + replacement_text + converted_text[end_pos:]

    # Generate output txt file (converted prices)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_txt_path = rf"{OUTPUTS_DIR}\Multi_converted_{timestamp}.txt"

    # Write converted text (CHF prices converted to EUR)
    with open(output_txt_path, 'w', encoding='utf-8') as f:
        f.write(converted_text)

    print(f"\n[OK] Output txt file saved: {output_txt_path}")

    # Generate filtered Stock Lines Excel
    if recognized_items:
        filtered_stock_rows = []
        item_nos = set()

        for item in recognized_items:
            item_no = item['item_no']
            if item_no not in item_nos:
                item_nos.add(item_no)
                filtered_stock_rows.append(item['stock_row'])

        if filtered_stock_rows:
            filtered_df = pd.DataFrame(filtered_stock_rows)
            output_excel_path = rf"{RECOGNITION_REPORT_DIR}\Stock_Lines_Filtered_{timestamp}.xlsx"
            filtered_df.to_excel(output_excel_path, index=False)
            print(f"[OK] Filtered Stock Lines Excel saved: {output_excel_path}")
            print(f"   Contains {len(filtered_df)} recognized items")

    # Generate Lines.xlsx by copying template and filling data
    # This preserves exact Excel formatting required by Business Central
    if recognized_items:
        try:
            # Create Main offer directory if it doesn't exist
            Path(MAIN_OFFER_DIR).mkdir(parents=True, exist_ok=True)
            lines_excel_path = rf"{MAIN_OFFER_DIR}\Lines.xlsx"

            # Copy template to preserve formatting
            copy2(LINES_TEMPLATE_PATH, lines_excel_path)

            # Open the copied template
            wb = openpyxl.load_workbook(lines_excel_path)
            ws = wb.active

            # Clear existing data (keep header row 1)
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)

            # Fill in matched wines starting from row 2
            for idx, item in enumerate(recognized_items, start=2):
                stock_row = item['stock_row']
                omt_row = item['omt_row']

                ws.cell(idx, 1).value = stock_row.get('Wine Name')  # Wine Name
                ws.cell(idx, 2).value = stock_row.get('Vintage Code')  # Vintage Code
                ws.cell(idx, 3).value = stock_row.get('Size')  # Size
                ws.cell(idx, 4).value = stock_row.get('Producer Name')  # Producer Name
                ws.cell(idx, 5).value = omt_row.get('Minimum Quantity')  # Minimum Quantity
                ws.cell(idx, 6).value = item['chf_price']  # Unit Price (CHF)
                ws.cell(idx, 7).value = stock_row.get('Sales Price Base Inc. VAT (CHF)')  # Unit Price Incl. VAT
                ws.cell(idx, 8).value = int(round(float(item['eur_price'])))  # Unit Price (€)
                ws.cell(idx, 9).value = omt_row.get('Main Offer Comment')  # Main Offer Comment
                ws.cell(idx, 10).value = None  # Competitor Code
                ws.cell(idx, 11).value = None  # Group Code

            # Save the workbook
            wb.save(lines_excel_path)
            wb.close()

            print(f"[OK] Lines.xlsx saved with {len(recognized_items)} recognized wines")
            print(f"   Saved to: {lines_excel_path}")
        except Exception as e:
            print(f"[WARN] Error generating Lines.xlsx: {e}")
            import traceback
            traceback.print_exc()

    # Generate recognition report in Detailed match results folder
    # Create directory if it doesn't exist
    Path(RECOGNITION_REPORT_DIR).mkdir(parents=True, exist_ok=True)
    report_path = rf"{RECOGNITION_REPORT_DIR}\Recognition_Report_{timestamp}.txt"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("WINE RECOGNITION REPORT\n")
        f.write("="*80 + "\n\n")
        f.write(f"Total prices found: {len(all_matches)}\n")
        f.write(f"Successfully matched: {len(recognized_items)}\n")
        f.write(f"Not matched: {len(all_matches) - len(recognized_items)}\n\n")
        f.write("="*80 + "\n")
        f.write("RECOGNIZED WINES:\n")
        f.write("="*80 + "\n\n")

        for item in recognized_items:
            omt_row = item['omt_row']
            stock_row = item['stock_row']
            # Use extracted wine name from Multi.txt, fallback to database name
            wine_name_display = item.get('extracted_wine_name') or omt_row.get('Wine Name')
            vintage_display = item.get('extracted_vintage') or omt_row.get('Vintage Code')

            f.write(f"Wine (from Multi.txt): {wine_name_display}\n")
            f.write(f"Vintage: {vintage_display}\n")
            f.write(f"Database Match: {stock_row.get('Wine Name')}\n")
            f.write(f"CHF Price: {item['chf_price']}\n")
            f.write(f"EUR Price: {item['eur_price']}\n")
            f.write(f"Item No.: {item['item_no']}\n")
            f.write(f"Size: {omt_row.get('Size')}\n")
            f.write(f"Producer: {omt_row.get('Producer Name')}\n")
            f.write("-"*80 + "\n")

    print(f"[OK] Recognition report saved: {report_path}")

    # Statistics
    print("\n" + "="*80)
    print("CONVERSION STATISTICS")
    print("="*80)
    print(f"Total CHF prices found: {len(all_matches)}")
    print(f"[OK] Successfully matched to Stock Lines: {len(recognized_items)}")
    print(f"[WARN] Not matched: {len(all_matches) - len(recognized_items)}")
    print("="*80 + "\n")

    # Generate translations using DeepL API (if enabled)
    if enable_translations:
        print("\n" + "="*80)
        print("GENERATING TRANSLATIONS")
        print("="*80 + "\n")

        # Create translations directory if it doesn't exist
        Path(TRANSLATIONS_DIR).mkdir(parents=True, exist_ok=True)

        # Translate to German
        print("[INFO] Translating to German...")
        german_text = translate_text_deepl(converted_text, 'DE')
        if german_text:
            german_output_path = rf"{TRANSLATIONS_DIR}\Multi_DE_{timestamp}.txt"
            with open(german_output_path, 'w', encoding='utf-8') as f:
                f.write(german_text)
            print(f"[OK] German translation saved: {german_output_path}")
        else:
            print("[WARN] German translation failed")

        # Translate to French
        print("[INFO] Translating to French...")
        french_text = translate_text_deepl(converted_text, 'FR')
        if french_text:
            french_output_path = rf"{TRANSLATIONS_DIR}\Multi_FR_{timestamp}.txt"
            with open(french_output_path, 'w', encoding='utf-8') as f:
                f.write(french_text)
            print(f"[OK] French translation saved: {french_output_path}")
        else:
            print("[WARN] French translation failed")

        print("\n" + "="*80)
        print("CONVERSION AND TRANSLATION COMPLETE")
        print("="*80 + "\n")
    else:
        print("\n[INFO] Translations skipped (checkbox disabled)")


if __name__ == "__main__":
    import sys
    # Check if translations should be enabled (default True)
    enable_trans = True
    if len(sys.argv) > 1:
        enable_trans = sys.argv[1] == "1"
    convert_txt_file(enable_translations=enable_trans)
