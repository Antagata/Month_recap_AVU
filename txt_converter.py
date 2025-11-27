#!/usr/bin/env python3
"""
FAST Text File CHF to EUR Converter
Optimized for large files with 100+ CHF prices
"""

import re
import pandas as pd
from datetime import datetime
from pathlib import Path
import math
import openpyxl
from shutil import copy2
import requests

# Configuration
BASE_DIR = r"C:\Users\Marco.Africani\Desktop\Month recap"
DATABASE_DIR = r"C:\Users\Marco.Africani\OneDrive - AVU SA\AVU CPI Campaign\Puzzle_control_Reports\SOURCE_FILES"
INPUT_FILE_PATH = rf"{BASE_DIR}\Inputs\Multi.txt"
STOCK_FILE_PATH = rf"{DATABASE_DIR}\Stock Lines.xlsx"
OMT_FILE_PATH = rf"{DATABASE_DIR}\OMT Main Offer List.xlsx"
LEARNING_DB_PATH = rf"{BASE_DIR}\wine_names_learning_db.txt"
OUTPUTS_DIR = rf"{BASE_DIR}\Outputs"
RECOGNITION_REPORT_DIR = rf"{BASE_DIR}\Outputs\Detailed match results"
OMT_LINES_DIR = rf"{BASE_DIR}\Outputs\OMT lines"
TRANSLATIONS_DIR = rf"{BASE_DIR}\Outputs\EUR translations"
MAIN_OFFER_DIR = rf"{BASE_DIR}\Outputs\Detailed match results\Main offer"
LINES_TEMPLATE_PATH = rf"{MAIN_OFFER_DIR}\template\Lines Template.xlsx"

# DeepL API Configuration
DEEPL_API_KEY = "374a8965-101a-4538-bc65-54506552650e"
DEEPL_API_URL = "https://api.deepl.com/v2/translate"

# CHF price pattern - handles: "CHF 900.00", "230.00 CHF", "29.00 CH"
CHF_PRICE_PATTERN = r'(?:\bCHF\s+(\d+(?:[\']\d{3})*(?:\.\d{2})?)|(\d+(?:[\']\d{3})*(?:\.\d{2})?)\s+CHF\b|(\d+(?:[\']\d{3})*(?:\.\d{2})?)\s+CH\s)'


def load_learning_database():
    """Load the learning database with wine name -> Item No. mappings"""
    learning_db = {}

    if not Path(LEARNING_DB_PATH).exists():
        print(f"[WARN] Learning database not found at {LEARNING_DB_PATH}")
        return learning_db

    try:
        with open(LEARNING_DB_PATH, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        for line in lines:
            line = line.strip()
            if line and not line.startswith('#'):
                parts = line.split(' | ')
                if len(parts) >= 3:
                    wine_name = parts[0].strip()
                    vintage = parts[1].strip()
                    item_no = parts[2].split()[0].strip()

                    if item_no != 'NOT_FOUND' and item_no.isdigit():
                        # Normalize wine name for lookup
                        clean_name = wine_name.lower().strip()
                        clean_name = re.sub(r'^(château|chateau|domaine|dom\.|ch\.)\s+', '', clean_name, flags=re.IGNORECASE)
                        clean_name = clean_name.replace('-', ' ')

                        # Use (clean_name, vintage) as key
                        key = (clean_name, vintage)
                        learning_db[key] = int(item_no)

        print(f"[OK] Loaded {len(learning_db)} wine mappings from learning database")
        return learning_db

    except Exception as e:
        print(f"[WARN] Error loading learning database: {e}")
        return {}


def load_databases():
    """Load Stock Lines and OMT databases once at startup"""
    print("Loading Stock Lines database...")
    stock_df = pd.read_excel(STOCK_FILE_PATH)
    print(f"[OK] Loaded {len(stock_df)} items from Stock Lines.xlsx")

    print("Loading OMT Main Offer List...")
    omt_df = pd.read_excel(OMT_FILE_PATH)
    print(f"[OK] Loaded {len(omt_df)} rows from OMT Main Offer List.xlsx")

    # Pre-process for faster lookups
    omt_df['Item No. Int'] = pd.to_numeric(omt_df['Item No.'], errors='coerce').astype('Int64')
    omt_df['Schedule DateTime'] = pd.to_datetime(omt_df['Schedule DateTime'])

    # Load learning database
    learning_db = load_learning_database()

    return stock_df, omt_df, learning_db


def extract_wine_name_vintage(text_before_price):
    """
    Extract wine name and vintage from text before price

    Special handling for Champagne wines:
    - May not have a vintage (NV or no year mentioned)
    - Producer names may appear in the wine name (e.g., "Krug Rosé 29ème Édition")
    - Detect "Magnum" to distinguish bottle sizes
    """
    # Look for pattern at start of line: "Wine Name VINTAGE :" or "Wine Name VINTAGE -"
    # First try to find line breaks to get the current line
    lines = text_before_price[-300:].split('\n')
    current_line = lines[-1] if lines else text_before_price[-150:]

    # Check if this is a Magnum (150cl bottle)
    is_magnum = 'magnum' in current_line.lower()

    # Pattern 1: "Wine Name VINTAGE : description" (most common)
    pattern1 = r'([A-Z\u00C0-\u017F][^\n\d]*?)\s+(19\d{2}|20\d{2}|NV)\s*:'
    match = re.search(pattern1, current_line)

    if match:
        wine_name = match.group(1).strip()
        vintage_str = match.group(2)
        vintage = int(vintage_str) if vintage_str != 'NV' else 0

        # Clean up emojis and special chars
        wine_name = re.sub(r'[✨💎💼🍷🏆⭐🎯]', '', wine_name).strip()

        # Add magnum indicator to wine name if present
        if is_magnum and 'magnum' not in wine_name.lower():
            wine_name = wine_name + ' Magnum'

        return wine_name, vintage

    # Pattern 2: "Wine Name VINTAGE -" or just "Wine Name VINTAGE"
    pattern2 = r'([A-Z\u00C0-\u017F][^\n\d]*?)\s+(19\d{2}|20\d{2}|NV)\s*[-:]?'
    match = re.search(pattern2, current_line)

    if match:
        wine_name = match.group(1).strip()
        vintage_str = match.group(2)
        vintage = int(vintage_str) if vintage_str != 'NV' else 0

        # Clean up emojis and special chars
        wine_name = re.sub(r'[✨💎💼🍷🏆⭐🎯]', '', wine_name).strip()

        # Add magnum indicator to wine name if present
        if is_magnum and 'magnum' not in wine_name.lower():
            wine_name = wine_name + ' Magnum'

        return wine_name, vintage

    # Pattern 3: Champagne without vintage (e.g., "Krug Rosé 29ème Édition :")
    # Match wine names that end with : but have no year
    pattern3 = r'([A-Z\u00C0-\u017F][^\n:]*?)\s*:'
    match = re.search(pattern3, current_line)

    if match:
        wine_name = match.group(1).strip()

        # Clean up emojis and special chars
        wine_name = re.sub(r'[✨💎💼🍷🏆⭐🎯]', '', wine_name).strip()

        # Add magnum indicator to wine name if present
        if is_magnum and 'magnum' not in wine_name.lower():
            wine_name = wine_name + ' Magnum'

        # For Champagne without vintage, return NV (0)
        return wine_name, 0

    return None, None


def match_chf_to_eur(chf_price, wine_name, vintage, stock_df, omt_df, learning_db, min_quantity=0, size_filter=None):
    """
    IMPROVED: Use learning database first, then try wine name + vintage + price match, then price-only fallback

    Args:
        size_filter: Optional size filter (e.g., 75.0 for standard, 150.0 for Magnum)
    """
    try:
        chf_float = float(chf_price)

        # Step 0: Check learning database first (if wine_name and vintage are available)
        if wine_name and vintage and learning_db:
            # Normalize wine name for learning DB lookup
            clean_name = wine_name.lower().strip()
            clean_name = re.sub(r'^(château|chateau|domaine|dom\.|ch\.)\s+', '', clean_name, flags=re.IGNORECASE)
            clean_name = clean_name.replace('-', ' ')

            vintage_str = str(vintage) if vintage != 0 else 'NV'
            lookup_key = (clean_name, vintage_str)

            if lookup_key in learning_db:
                learned_item_no = learning_db[lookup_key]

                # Find this item in Stock Lines
                stock_match = stock_df[stock_df['No.'] == learned_item_no]

                if len(stock_match) > 0:
                    stock_row = stock_match.iloc[0]

                    # Find in OMT with the learned Item No. + price + min_quantity
                    # FILTERS: Competitor Code must be empty, Campaign Status must be 'Sent'
                    omt_matches = omt_df[
                        (omt_df['Item No. Int'] == learned_item_no) &
                        (omt_df['Unit Price'].astype(float).round(2) == round(chf_float, 2)) &
                        (omt_df['Minimum Quantity'].astype(float) == float(min_quantity)) &
                        (omt_df['Campaign Sub-Type'] == 'Normal') &
                        (omt_df['Campaign Type'] == 'PRIVATE') &
                        (omt_df['Competitor Code'].isna() | (omt_df['Competitor Code'] == '')) &  # Rule 1: Competitor Code empty
                        (omt_df['Campaign Status'] == 'Sent')  # Rule 2: Campaign Status = Sent
                    ]

                    if len(omt_matches) > 0:
                        omt_row = omt_matches.iloc[0]
                        return {
                            'item_no': learned_item_no,
                            'chf_price': chf_price,
                            'eur_price': omt_row.get('Unit Price (EUR)', 0),
                            'stock_row': stock_row,
                            'omt_row': omt_row,
                            'schedule_datetime': omt_row['Schedule DateTime'],
                            'wine_name': wine_name,
                            'vintage': vintage,
                            'matched_by_name': True,  # Matched from learning DB
                            'from_learning_db': True
                        }

        # Step 1: Find in Stock Lines by OMT Last Private Offer Price
        stock_matches = stock_df[
            stock_df['OMT Last Private Offer Price'].astype(float).round(2) == round(chf_float, 2)
        ]

        # Apply size filter if specified (Rule 3: Filter by bottle size for Magnums)
        if size_filter is not None:
            stock_matches = stock_matches[stock_matches['Size'].astype(float) == float(size_filter)]

        if len(stock_matches) == 0:
            return None

        # Step 1.5: If we have wine_name and vintage, try to filter stock_matches by name/vintage first
        name_filtered_matches = stock_matches
        used_name_filter = False  # Track if we successfully filtered by name/vintage

        if wine_name and vintage:
            # Try to match by vintage first (most reliable)
            # NOTE: Vintage Code in database is STRING, so convert to string for comparison
            vintage_str = str(vintage) if vintage != 0 else 'NV'
            vintage_matches = stock_matches[stock_matches['Vintage Code'].astype(str) == vintage_str]
            if len(vintage_matches) > 0:
                # Then try to match by wine name (fuzzy)
                # Clean wine name for comparison - normalize hyphens and spaces
                clean_name = wine_name.lower().strip()
                # Remove common prefixes
                clean_name = re.sub(r'^(château|chateau|domaine|dom\.|ch\.)\s+', '', clean_name, flags=re.IGNORECASE)
                # Normalize: replace hyphens with spaces for comparison
                clean_name_normalized = clean_name.replace('-', ' ')

                # Try to find wines with matching name parts
                name_matches = []
                for _, row in vintage_matches.iterrows():
                    db_wine_name = str(row.get('Wine Name', '')).lower()
                    # Normalize database name too
                    db_wine_normalized = db_wine_name.replace('-', ' ')

                    # Check if significant parts of the name appear in the database wine name
                    # Split by spaces and take first 3 significant words
                    name_parts = [p for p in clean_name_normalized.split() if len(p) > 3][:3]

                    if name_parts and any(part in db_wine_normalized for part in name_parts):
                        name_matches.append(row)
                        used_name_filter = True  # We found a name match!

                if len(name_matches) > 0:
                    name_filtered_matches = pd.DataFrame(name_matches)
                else:
                    # If no name matches but we have vintage matches, use vintage matches
                    name_filtered_matches = vintage_matches
                    # Don't set used_name_filter=True because we only matched by vintage, not name

        # Step 2: For each Stock Lines match, find in OMT
        all_candidates = []

        for _, stock_row in name_filtered_matches.iterrows():
            item_no = stock_row['No.']

            # BULLETPROOF MATCH: Item No. + Unit Price (CHF) + Minimum Quantity
            # FILTERS: Campaign Sub-Type = Normal, Campaign Type = PRIVATE,
            #          Competitor Code empty, Campaign Status = Sent
            omt_matches = omt_df[
                (omt_df['Item No. Int'] == item_no) &
                (omt_df['Unit Price'].astype(float).round(2) == round(chf_float, 2)) &
                (omt_df['Minimum Quantity'].astype(float) == float(min_quantity)) &
                (omt_df['Campaign Sub-Type'] == 'Normal') &
                (omt_df['Campaign Type'] == 'PRIVATE') &
                (omt_df['Competitor Code'].isna() | (omt_df['Competitor Code'] == '')) &  # Rule 1: Competitor Code empty
                (omt_df['Campaign Status'] == 'Sent')  # Rule 2: Campaign Status = Sent
            ]

            for _, omt_row in omt_matches.iterrows():
                all_candidates.append({
                    'item_no': item_no,
                    'chf_price': chf_price,
                    'eur_price': omt_row.get('Unit Price (EUR)', 0),
                    'stock_row': stock_row,
                    'omt_row': omt_row,
                    'schedule_datetime': omt_row['Schedule DateTime'],
                    'wine_name': wine_name,
                    'vintage': vintage,
                    'matched_by_name': used_name_filter  # True if we used name filtering
                })

        if len(all_candidates) == 0:
            return None

        # Step 3: Pick latest Schedule DateTime
        all_candidates.sort(key=lambda x: x['schedule_datetime'], reverse=True)
        return all_candidates[0]

    except Exception as e:
        print(f"[WARN] Error matching CHF {chf_price}: {e}")
        return None


def translate_text_deepl(text, target_language):
    """Translate text using DeepL API"""
    try:
        params = {
            'auth_key': DEEPL_API_KEY,
            'text': text,
            'target_lang': target_language
        }
        response = requests.post(DEEPL_API_URL, data=params, timeout=30)

        if response.status_code == 200:
            return response.json()['translations'][0]['text']
        else:
            print(f"[WARN] DeepL API error {response.status_code}: {response.text}")
            return None
    except Exception as e:
        print(f"[WARN] Translation error: {e}")
        return None


def convert_txt_file_fast(enable_translations=True):
    """Fast CHF to EUR conversion for large files"""

    print("="*80)
    print("TXT File CHF to EUR Converter (FAST)")
    print("="*80)
    print()

    # Load databases once
    stock_df, omt_df, learning_db = load_databases()

    # Load input file
    print(f"[OK] Loaded input file: {INPUT_FILE_PATH}")
    with open(INPUT_FILE_PATH, 'r', encoding='utf-8') as f:
        text = f.read()

    # Find all CHF prices
    chf_matches = list(re.finditer(CHF_PRICE_PATTERN, text))
    print(f"[INFO] Found {len(chf_matches)} CHF prices in input file")
    print()

    # Process all prices
    recognized_items = []
    matched_omt_rows = []  # For OMT Excel export
    replacements = []  # (start, end, replacement_text)
    processed_positions = set()  # Track which prices we've already processed
    corrections_needed = []  # Track wines that need manual correction

    for i, match in enumerate(chf_matches, 1):
        # Skip if already processed (for paired prices)
        if match.start() in processed_positions:
            continue

        # Extract price value
        price_str_raw = match.group(1) or match.group(2) or match.group(3)
        if not price_str_raw:
            continue

        # Remove thousands separator (')
        price_str = price_str_raw.replace("'", "")
        price_value = float(price_str)

        # Get text before price for wine name extraction
        text_before = text[max(0, match.start()-200):match.start()]
        wine_name, vintage = extract_wine_name_vintage(text_before)

        # Check if there's a second price on the same line for 36+ bottles
        # Look ahead up to 100 chars for patterns like "// 290.00 CHF for 36+" or "36x 24.00 CHF"
        text_after = text[match.end():min(match.end()+100, len(text))]
        second_price_match = None
        is_36_bottle_price = False

        # Look for "// PRICE" or "36x PRICE" or "36+ bottles" indicators
        if re.search(r'//', text_after):
            # Find the next CHF price after "//"
            for next_match in chf_matches[i:]:  # Start from current position
                if next_match.start() > match.end():
                    between_text = text[match.end():next_match.start()]
                    if '//' in between_text and re.search(r'(36|for\s+36)', between_text, re.IGNORECASE):
                        second_price_match = next_match
                        is_36_bottle_price = True
                        processed_positions.add(next_match.start())
                        break

        # Detect bottle size: Check if "Magnum" is in wine name or text before price
        size_filter = None
        if wine_name and 'magnum' in wine_name.lower():
            size_filter = 150.0  # Magnum size
        else:
            size_filter = 75.0   # Standard bottle size

        # Match the first price (min_quantity=0)
        item_info = match_chf_to_eur(price_value, wine_name, vintage, stock_df, omt_df, learning_db, min_quantity=0, size_filter=size_filter)

        if item_info:
            # Add order index to preserve original position from Multi.txt
            item_info['order_index'] = i  # i is the position in the loop (1-indexed)
            recognized_items.append(item_info)
            matched_omt_rows.append(item_info['omt_row'])

            # Check if this match is questionable (price-only match without name confirmation)
            # Only flag if we have a wine_name but it WASN'T matched by name
            if wine_name and not item_info.get('matched_by_name', False):
                stock_wine = item_info['stock_row'].get('Wine Name', '')
                stock_vintage = item_info['stock_row'].get('Vintage Code', '')
                corrections_needed.append({
                    'wine_from_multi': wine_name,
                    'vintage_from_multi': vintage,
                    'chf_price': price_value,
                    'matched_wine': stock_wine,
                    'matched_vintage': stock_vintage,
                    'item_no': item_info['item_no'],
                    'reason': 'Price-only match - wine name did not match database',
                    'min_qty': 0
                })

            # Now match the second price (min_quantity=36) if it exists
            if second_price_match:
                second_price_str_raw = second_price_match.group(1) or second_price_match.group(2) or second_price_match.group(3)
                if second_price_str_raw:
                    second_price_str = second_price_str_raw.replace("'", "")
                    second_price_value = float(second_price_str)

                    # Use the SAME item_no from the first match, but look for the 36-bottle price
                    item_no = item_info['item_no']
                    stock_row = item_info['stock_row']

                    # Find in OMT with min_quantity=36
                    omt_matches_36 = omt_df[
                        (omt_df['Item No. Int'] == item_no) &
                        (omt_df['Unit Price'].astype(float).round(2) == round(second_price_value, 2)) &
                        (omt_df['Minimum Quantity'].astype(float) == 36.0) &
                        (omt_df['Campaign Sub-Type'] == 'Normal') &
                        (omt_df['Campaign Type'] == 'PRIVATE') &
                        (omt_df['Competitor Code'].isna() | (omt_df['Competitor Code'] == '')) &  # Rule 1
                        (omt_df['Campaign Status'] == 'Sent')  # Rule 2
                    ]

                    if len(omt_matches_36) > 0:
                        omt_row_36 = omt_matches_36.iloc[0]
                        item_info_36 = {
                            'item_no': item_no,
                            'chf_price': second_price_value,
                            'eur_price': omt_row_36.get('Unit Price (EUR)', 0),
                            'stock_row': stock_row,
                            'omt_row': omt_row_36,
                            'schedule_datetime': omt_row_36['Schedule DateTime'],
                            'wine_name': wine_name,
                            'vintage': vintage,
                            'order_index': i  # Same order as parent wine
                        }
                        recognized_items.append(item_info_36)
                        matched_omt_rows.append(omt_row_36)

                        # Add EUR replacement for second price
                        eur_price_36 = item_info_36['eur_price']
                        eur_rounded_36 = f"{int(round(float(eur_price_36)))}.00"

                        original_text_36 = second_price_match.group(0)
                        if 'CHF' in original_text_36:
                            eur_text_36 = original_text_36.replace('CHF', 'EUR')
                            eur_text_36 = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded_36, eur_text_36)
                        elif 'CH ' in original_text_36:
                            eur_text_36 = original_text_36.replace('CH ', 'EUR ')
                            eur_text_36 = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded_36, eur_text_36)
                        else:
                            eur_text_36 = original_text_36.replace(second_price_str_raw, eur_rounded_36)

                        replacements.append((second_price_match.start(), second_price_match.end(), eur_text_36))
                        print(f"[{i}/{len(chf_matches)}] Matched 36-bottle: {wine_name} {vintage} - CHF {second_price_value} -> EUR {eur_price_36} (Item {item_no}, Min Qty 36)")

            # Add EUR replacement for first price
            eur_price = item_info['eur_price']
            eur_rounded = f"{int(round(float(eur_price)))}.00"

            original_text = match.group(0)
            if 'CHF' in original_text:
                eur_text = original_text.replace('CHF', 'EUR')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            elif 'CH ' in original_text:
                eur_text = original_text.replace('CH ', 'EUR ')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            else:
                eur_text = original_text.replace(price_str_raw, eur_rounded)

            replacements.append((match.start(), match.end(), eur_text))

            safe_name = (wine_name or '').encode('ascii', errors='ignore').decode('ascii')
            matched_by_name_flag = item_info.get('matched_by_name', False)
            from_learning_db = item_info.get('from_learning_db', False)

            if from_learning_db:
                match_type = "LEARNING-DB"
            elif matched_by_name_flag:
                match_type = "NAME+VINTAGE"
            else:
                match_type = "PRICE-ONLY"

            print(f"[{i}/{len(chf_matches)}] Matched ({match_type}): {safe_name} {vintage} - CHF {price_value} -> EUR {eur_price} (Item {item_info['item_no']})")
        else:
            # No match found - add to corrections needed
            corrections_needed.append({
                'wine_from_multi': wine_name,
                'vintage_from_multi': vintage,
                'chf_price': price_value,
                'matched_wine': 'NO MATCH FOUND',
                'matched_vintage': '',
                'item_no': 'MANUAL_ENTRY_NEEDED',
                'reason': 'No database match found for this wine at this price',
                'min_qty': 0
            })

            # Try fallback for unmatched first price
            eur_fallback = price_value * 1.08
            eur_rounded = f"{int(math.floor(eur_fallback))}.00"

            original_text = match.group(0)
            if 'CHF' in original_text:
                eur_text = original_text.replace('CHF', 'EUR')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            else:
                eur_text = original_text.replace(price_str_raw, eur_rounded)

            replacements.append((match.start(), match.end(), eur_text))

            safe_name = (wine_name or '').encode('ascii', errors='ignore').decode('ascii')
            print(f"[{i}/{len(chf_matches)}] Fallback: {safe_name} {vintage} - CHF {price_value} -> EUR {eur_fallback:.2f} (no match)")

            # Also handle fallback for second price if it exists
            if second_price_match:
                second_price_str_raw = second_price_match.group(1) or second_price_match.group(2) or second_price_match.group(3)
                if second_price_str_raw:
                    second_price_str = second_price_str_raw.replace("'", "")
                    second_price_value = float(second_price_str)
                    eur_fallback_36 = second_price_value * 1.08
                    eur_rounded_36 = f"{int(math.floor(eur_fallback_36))}.00"

                    original_text_36 = second_price_match.group(0)
                    if 'CHF' in original_text_36:
                        eur_text_36 = original_text_36.replace('CHF', 'EUR')
                        eur_text_36 = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded_36, eur_text_36)
                    else:
                        eur_text_36 = original_text_36.replace(second_price_str_raw, eur_rounded_36)

                    replacements.append((second_price_match.start(), second_price_match.end(), eur_text_36))
                    print(f"[{i}/{len(chf_matches)}] Fallback 36-bottle: {wine_name} {vintage} - CHF {second_price_value} -> EUR {eur_fallback_36:.2f} (no match)")

    # Apply all replacements in reverse order
    converted_text = text
    replacements.sort(key=lambda x: x[0], reverse=True)
    for start, end, replacement in replacements:
        converted_text = converted_text[:start] + replacement + converted_text[end:]

    # Generate timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Save converted text
    output_txt_path = rf"{OUTPUTS_DIR}\Multi_converted_{timestamp}.txt"
    with open(output_txt_path, 'w', encoding='utf-8') as f:
        f.write(converted_text)
    print(f"\n[OK] Output txt file saved: {output_txt_path}")

    # Generate CORRECTIONS_NEEDED file if there are questionable matches
    if corrections_needed:
        corrections_path = rf"{RECOGNITION_REPORT_DIR}\CORRECTIONS_NEEDED_{timestamp}.txt"
        with open(corrections_path, 'w', encoding='utf-8') as f:
            f.write("="*120 + "\n")
            f.write("WINE MATCHING CORRECTIONS NEEDED\n")
            f.write("="*120 + "\n")
            f.write("\nThis file lists wines that either:\n")
            f.write("1. Could not be matched in the database\n")
            f.write("2. Were matched by PRICE ONLY (wine name could not be verified)\n")
            f.write("\nHOW TO MAKE CORRECTIONS:\n")
            f.write("- For each wine below, check if the matched Item Number is correct\n")
            f.write("- If WRONG, enter the correct Item Number in the 'CORRECTED_ITEM_NO' field\n")
            f.write("- Leave the field empty if the match is correct\n")
            f.write("- After making corrections, click 'Apply Corrections' in the GUI\n")
            f.write("\nFORMAT: CORRECTED_ITEM_NO: [enter number here or leave empty]\n")
            f.write("="*120 + "\n\n")

            for idx, correction in enumerate(corrections_needed, 1):
                f.write(f"\n{idx}. WINE FROM MULTI.TXT:\n")
                f.write(f"   Name: {correction['wine_from_multi'] or 'Unknown'}\n")
                f.write(f"   Vintage: {correction['vintage_from_multi'] or 'Unknown'}\n")
                f.write(f"   CHF Price: {correction['chf_price']}\n")
                f.write(f"   Min Qty: {correction['min_qty']}\n")
                f.write(f"\n   MATCHED TO DATABASE:\n")
                f.write(f"   Wine: {correction['matched_wine']}\n")
                f.write(f"   Vintage: {correction['matched_vintage']}\n")
                f.write(f"   Item No.: {correction['item_no']}\n")
                f.write(f"\n   REASON: {correction['reason']}\n")
                f.write(f"\n   >>> CORRECTED_ITEM_NO: \n")  # Manual entry field
                f.write(f"   {'='*100}\n")

        print(f"[WARN] Generated corrections file with {len(corrections_needed)} questionable matches")
        print(f"       Please review: {corrections_path}")
    else:
        print(f"\n[OK] All wines matched with high confidence (name + vintage + price)")

    # Generate Lines.xlsx ONLY (all other Excel files removed)
    if recognized_items:
        try:
            Path(MAIN_OFFER_DIR).mkdir(parents=True, exist_ok=True)
            # Add timestamp to Lines.xlsx filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            lines_excel_path = rf"{MAIN_OFFER_DIR}\Lines_{timestamp}.xlsx"

            # Copy template
            copy2(LINES_TEMPLATE_PATH, lines_excel_path)

            # Fill data
            wb = openpyxl.load_workbook(lines_excel_path)
            ws = wb.active

            # Clear existing data rows (keep header)
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)

            # IMPORTANT: Sort recognized_items by order_index to preserve Multi.txt order
            recognized_items_sorted = sorted(recognized_items, key=lambda x: x.get('order_index', 0))

            # Write recognized items (in the exact order from Multi.txt)
            for idx, item in enumerate(recognized_items_sorted, start=2):
                stock_row = item['stock_row']
                omt_row = item['omt_row']

                ws.cell(idx, 1).value = stock_row.get('Wine Name')
                ws.cell(idx, 2).value = stock_row.get('Vintage Code')
                ws.cell(idx, 3).value = stock_row.get('Size')
                ws.cell(idx, 4).value = stock_row.get('Producer Name')
                ws.cell(idx, 5).value = omt_row.get('Minimum Quantity')
                ws.cell(idx, 6).value = item['chf_price']
                ws.cell(idx, 7).value = stock_row.get('Sales Price Base Inc. VAT (CHF)')
                ws.cell(idx, 8).value = int(round(float(item['eur_price'])))  # Rounded EUR
                ws.cell(idx, 9).value = omt_row.get('Main Offer Comment')
                ws.cell(idx, 10).value = None
                ws.cell(idx, 11).value = None
                ws.cell(idx, 12).value = item['item_no']  # Item Number in column L

            wb.save(lines_excel_path)
            wb.close()

            print(f"\n[OK] Lines_{timestamp}.xlsx saved with {len(recognized_items)} wines")
            print(f"   Saved to: {lines_excel_path}")
            print(f"   Wines are listed in the same order as Multi.txt")
            print(f"   Item Numbers are in column L (12)")
        except Exception as e:
            print(f"[ERROR] Failed to generate Lines_{timestamp}.xlsx: {e}")
    else:
        print("\n[WARN] No wines were recognized, Lines.xlsx not generated")

    # Statistics
    print("\n" + "="*80)
    print("CONVERSION STATISTICS")
    print("="*80)
    print(f"Total CHF prices found: {len(chf_matches)}")
    print(f"[OK] Successfully matched to Stock Lines: {len(recognized_items)}")
    print(f"[WARN] Not matched: {len(chf_matches) - len(recognized_items)}")
    print("="*80)
    print()

    # Translations
    if enable_translations:
        print("="*80)
        print("GENERATING TRANSLATIONS")
        print("="*80)
        print()

        Path(TRANSLATIONS_DIR).mkdir(parents=True, exist_ok=True)

        # German
        print("[INFO] Translating to German...")
        de_text = translate_text_deepl(converted_text, 'DE')
        if de_text:
            de_path = rf"{TRANSLATIONS_DIR}\Multi_DE_{timestamp}.txt"
            with open(de_path, 'w', encoding='utf-8') as f:
                f.write(de_text)
            print(f"[OK] German translation saved: {de_path}")

        # French
        print("[INFO] Translating to French...")
        fr_text = translate_text_deepl(converted_text, 'FR')
        if fr_text:
            fr_path = rf"{TRANSLATIONS_DIR}\Multi_FR_{timestamp}.txt"
            with open(fr_path, 'w', encoding='utf-8') as f:
                f.write(fr_text)
            print(f"[OK] French translation saved: {fr_path}")

        print("\n" + "="*80)
        print("CONVERSION AND TRANSLATION COMPLETE")
        print("="*80)
    else:
        print("\n[INFO] Translations skipped (checkbox disabled)")


if __name__ == "__main__":
    import sys
    enable_trans = True
    if len(sys.argv) > 1:
        enable_trans = sys.argv[1] == "1"
    convert_txt_file_fast(enable_translations=enable_trans)
