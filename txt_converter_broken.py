#!/usr/bin/env python3
"""
FAST Text File CHF to EUR Converter
Optimized for large files with 100+ CHF prices
"""

import re
import pandas as pd
from datetime import datetime
from pathlib import Path
import math
import openpyxl
from shutil import copy2
import requests

# Configuration
BASE_DIR = r"C:\Users\Marco.Africani\Desktop\Month recap"
DATABASE_DIR = r"C:\Users\Marco.Africani\OneDrive - AVU SA\AVU CPI Campaign\Puzzle_control_Reports\SOURCE_FILES"
INPUT_FILE_PATH = rf"{BASE_DIR}\Inputs\Multi.txt"
STOCK_FILE_PATH = rf"{DATABASE_DIR}\Stock Lines.xlsx"
OMT_FILE_PATH = rf"{DATABASE_DIR}\OMT Main Offer List.xlsx"
OUTPUTS_DIR = rf"{BASE_DIR}\Outputs"
RECOGNITION_REPORT_DIR = rf"{BASE_DIR}\Outputs\Detailed match results"
OMT_LINES_DIR = rf"{BASE_DIR}\Outputs\OMT lines"
TRANSLATIONS_DIR = rf"{BASE_DIR}\Outputs\EUR translations"
MAIN_OFFER_DIR = rf"{BASE_DIR}\Outputs\Detailed match results\Main offer"
LINES_TEMPLATE_PATH = rf"{MAIN_OFFER_DIR}\template\Lines Template.xlsx"

# DeepL API Configuration
DEEPL_API_KEY = "374a8965-101a-4538-bc65-54506552650e"
DEEPL_API_URL = "https://api.deepl.com/v2/translate"

# CHF price pattern - handles: "CHF 900.00", "230.00 CHF", "29.00 CH"
CHF_PRICE_PATTERN = r'(?:\bCHF\s+(\d+(?:[\']\d{3})*(?:\.\d{2})?)|(\d+(?:[\']\d{3})*(?:\.\d{2})?)\s+CHF\b|(\d+(?:[\']\d{3})*(?:\.\d{2})?)\s+CH\s)'


def load_databases():
    """Load Stock Lines, OMT, and Learning databases once at startup"""
    print("Loading Stock Lines database...")
    stock_df = pd.read_excel(STOCK_FILE_PATH)
    print(f"[OK] Loaded {len(stock_df)} items from Stock Lines.xlsx")

    print("Loading OMT Main Offer List...")
    omt_df = pd.read_excel(OMT_FILE_PATH)
    print(f"[OK] Loaded {len(omt_df)} rows from OMT Main Offer List.xlsx")

    # Pre-process for faster lookups
    omt_df['Item No. Int'] = pd.to_numeric(omt_df['Item No.'], errors='coerce').astype('Int64')
    omt_df['Schedule DateTime'] = pd.to_datetime(omt_df['Schedule DateTime'])

    # Load learning database for manual mappings
    learning_db = {}
    learning_db_path = rf"{BASE_DIR}\wine_names_learning_db.txt"
    try:
        with open(learning_db_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    try:
                        parts = line.split('|')
                        if len(parts) >= 3:
                            wine_name = parts[0].strip()
                            vintage = parts[1].strip()
                            item_no = parts[2].strip()

                            # Skip if item number is not valid
                            if item_no and item_no != 'NOT_FOUND' and item_no.isdigit():
                                # Create key: "wine_name vintage"
                                key = f"{wine_name.lower()} {vintage}".strip()
                                learning_db[key] = int(item_no)
                    except:
                        pass
        print(f"[OK] Loaded {len(learning_db)} manual mappings from learning database")
    except FileNotFoundError:
        print("[INFO] No learning database found, will create on first correction")

    return stock_df, omt_df, learning_db


def extract_wine_name_vintage(text_before_price, full_text, price_pos):
    """Extract wine name and vintage from text before price"""
    # Get the full line containing the price
    # Find line start by going backwards until newline
    line_start = full_text.rfind('\n', 0, price_pos) + 1
    line_end = full_text.find('\n', price_pos)
    if line_end == -1:
        line_end = len(full_text)

    line = full_text[line_start:line_end].strip()

    # Remove any leading bullet points, numbers, emojis, section markers
    line = re.sub(r'^[\d\.\)\-\*\u2022\u2728\u2713\u2714\u2716\u2717\u2718\u2719\u271a\u271b\u271c\u271d\u271e\u271f\u2720\u2721\u2722\u2723\u2724\u2725\u2726\u2727\u2728\u2729\u272a\u272b\u272c\u272d\u272e\u272f\u2730\u2731\u2732\u2733\u2734\u2735\u2736\u2737\u2738\u2739\u273a\u273b\u273c\u273d\u273e\u273f\u2740\u2741\u2742\u2743\u2744\u2745\u2746\u2747\u2748\u2749\u274a\u274b\u2b50\u2b51\u2b52\u2b53\u2b54\u2b55\u1f31f\u1f320\u1f4a1\u1f4a2\u1f4a3\u1f4a4\u1f4a5\u1f4a6\u1f4a7\u1f4a8\u1f4a9\u1f4aa\u1f4ab\u1f4ac\u1f4ad\u1f4ae\u1f4af\u1f4b0\u1f4b1\u1f4b2\u1f4b3\u1f4b4\u1f4b5\u1f4b6\u1f4b7\u1f4b8\u1f4b9\u1f4ba\u1f4bb\u1f4bc\u1f4bd\u1f4be\u1f4bf\u1f4c0\u1f575\u1f576\u1f5a4\u1f5a5\u1f681\u1f682\u1f683\u1f684\u1f685\u1f686\u1f687\u1f688\u1f689\u1f68a\u1f68b\u1f68c\u1f68d\u1f68e\u1f68f\u1f690\u1f691\u1f692\u1f693\u1f694\u1f695\u1f696\u1f697\u1f698\u1f699\u1f69a\u1f69b\u1f69c\u1f69d\u1f69e\u1f69f\u1f6a0\u1f6a1\u1f6a2\u1f6a3\u1f6a4\u1f6a5\u1f6a6\u1f6a7\u1f6a8\u1f6a9\u1f6aa\u1f6ab\u1f6ac\u1f6ad\u1f6ae\u1f6af\u1f6b0\u1f6b1\u1f6b2\u1f6b3\u1f6b4\u1f6b5\u1f6b6\u1f6b7\u1f6b8\u1f6b9\u1f6ba\u1f6bb\u1f6bc\u1f6bd\u1f6be\u1f6bf\u1f6c0\U0001f4bc\U0001f4b0\U0001f4b5\U0001f31f\U0001f575\U0001f5a4]\s*', '', line)

    # Skip section headers (lines that start with Top, CHF, etc.)
    if re.match(r'^(Top\s|CHF\s|MORE\s|Wines?\s)', line, re.IGNORECASE):
        return None, None

    # Remove price pattern from line first
    wine_line = re.sub(CHF_PRICE_PATTERN, '', line).strip()
    wine_line = re.sub(r'\s*\+\s*VAT\s*$', '', wine_line, flags=re.IGNORECASE).strip()

    # Split by " : " to get wine name part (before description)
    wine_part = wine_line.split(' : ')[0].strip()

    # If no colon, split by " - " for alternate format
    if wine_part == wine_line:
        # Check if there's a dash with description after it
        parts = wine_line.split(' – ')
        if len(parts) > 1:
            # Take only the first part
            wine_part = parts[0].strip()
        else:
            parts = wine_line.split(' - ')
            if len(parts) > 1:
                wine_part = parts[0].strip()
            else:
                wine_part = wine_line.strip()

    # Extract vintage if present (last 4-digit year in wine part)
    vintage_match = re.search(r'\b(19\d{2}|20\d{2})\b(?!.*\b(19\d{2}|20\d{2})\b)', wine_part)

    if vintage_match:
        vintage = vintage_match.group(1)
        # Remove vintage from wine name
        wine_name = wine_part[:vintage_match.start()].strip()
        # Remove trailing punctuation
        wine_name = re.sub(r'[:\-–,]\s*$', '', wine_name).strip()
        return wine_name, vintage
    else:
        # Remove trailing punctuation
        wine_name = re.sub(r'[:\-–,]\s*$', '', wine_part).strip()
        return wine_name, None


def match_chf_to_eur(chf_price, wine_name, vintage, stock_df, omt_df, learning_db, min_quantity=0):
    """
    CORRECTED MATCHING: Learning DB → Wine Name + Vintage + Price + Min Quantity
    """
    try:
        chf_float = float(chf_price)

        # Skip if no wine name extracted (section headers, etc.)
        if not wine_name or wine_name.strip() == '':
            return None

        # Step 1: Check learning database first for manual mappings
        learning_key = f"{wine_name.lower()} {vintage or ''}".strip()
        if learning_key in learning_db:
            item_no = learning_db[learning_key]
            print(f"[LEARNING] Found manual mapping for '{wine_name} {vintage}' → Item {item_no}")

            # Get from Stock Lines
            stock_row = stock_df[stock_df['No.'] == item_no]
            if len(stock_row) > 0:
                stock_row = stock_row.iloc[0]

                # Find in OMT
                omt_matches = omt_df[
                    (omt_df['Item No. Int'] == item_no) &
                    (omt_df['Unit Price'].astype(float).round(2) == round(chf_float, 2)) &
                    (omt_df['Minimum Quantity'].astype(float) == float(min_quantity)) &
                    (omt_df['Campaign Sub-Type'] == 'Normal') &
                    (omt_df['Campaign Type'] == 'PRIVATE')
                ]

                if len(omt_matches) > 0:
                    omt_row = omt_matches.iloc[0]
                    return {
                        'item_no': item_no,
                        'chf_price': chf_price,
                        'eur_price': omt_row.get('Unit Price (EUR)', 0),
                        'stock_row': stock_row,
                        'omt_row': omt_row,
                        'schedule_datetime': omt_row['Schedule DateTime'],
                        'wine_name': wine_name,
                        'vintage': vintage
                    }

        # Step 2: Find in Stock Lines by Wine Name (fuzzy matching)
        # Clean the extracted wine name
        wine_name_clean = wine_name.strip().lower()

        # Remove common prefixes for matching
        for prefix in ['château', 'chateau', 'domaine', 'dom.', 'ch.']:
            if wine_name_clean.startswith(prefix + ' '):
                wine_name_clean = wine_name_clean[len(prefix)+1:]
                break

        # Search Stock Lines for matching wine name
        stock_matches = []
        for _, stock_row in stock_df.iterrows():
            stock_wine = str(stock_row.get('Wine Name', '')).strip().lower()

            # Remove common prefixes from stock wine name too
            for prefix in ['château', 'chateau', 'domaine', 'dom.', 'ch.']:
                if stock_wine.startswith(prefix + ' '):
                    stock_wine = stock_wine[len(prefix)+1:]
                    break

            # Check if wine names match (fuzzy)
            if wine_name_clean in stock_wine or stock_wine in wine_name_clean:
                # Verify vintage matches if provided
                if vintage:
                    stock_vintage = str(stock_row.get('Vintage Code', '')).strip()
                    if stock_vintage != str(vintage):
                        continue

                # Verify price matches (within 1 CHF tolerance)
                stock_price = stock_row.get('OMT Last Private Offer Price', 0)
                try:
                    stock_price_float = float(stock_price)
                    if abs(stock_price_float - chf_float) <= 1.0:
                        stock_matches.append(stock_row)
                except:
                    pass

        if len(stock_matches) == 0:
            return None

        # Step 2: For each Stock Lines match, find in OMT
        all_candidates = []

        for _, stock_row in pd.DataFrame(stock_matches).iterrows():
            item_no = stock_row['No.']

            # BULLETPROOF MATCH: Item No. + Unit Price (CHF) + Minimum Quantity
            # FILTER: Campaign Sub-Type = Normal AND Campaign Type = PRIVATE
            omt_matches = omt_df[
                (omt_df['Item No. Int'] == item_no) &
                (omt_df['Unit Price'].astype(float).round(2) == round(chf_float, 2)) &
                (omt_df['Minimum Quantity'].astype(float) == float(min_quantity)) &
                (omt_df['Campaign Sub-Type'] == 'Normal') &
                (omt_df['Campaign Type'] == 'PRIVATE')
            ]

            for _, omt_row in omt_matches.iterrows():
                all_candidates.append({
                    'item_no': item_no,
                    'chf_price': chf_price,
                    'eur_price': omt_row.get('Unit Price (EUR)', 0),
                    'stock_row': stock_row,
                    'omt_row': omt_row,
                    'schedule_datetime': omt_row['Schedule DateTime'],
                    'wine_name': wine_name,
                    'vintage': vintage
                })

        if len(all_candidates) == 0:
            return None

        # Step 3: Pick latest Schedule DateTime
        all_candidates.sort(key=lambda x: x['schedule_datetime'], reverse=True)
        return all_candidates[0]

    except Exception as e:
        print(f"[WARN] Error matching wine '{wine_name}' CHF {chf_price}: {e}")
        return None


def translate_text_deepl(text, target_language):
    """Translate text using DeepL API"""
    try:
        params = {
            'auth_key': DEEPL_API_KEY,
            'text': text,
            'target_lang': target_language
        }
        response = requests.post(DEEPL_API_URL, data=params, timeout=30)

        if response.status_code == 200:
            return response.json()['translations'][0]['text']
        else:
            print(f"[WARN] DeepL API error {response.status_code}: {response.text}")
            return None
    except Exception as e:
        print(f"[WARN] Translation error: {e}")
        return None


def convert_txt_file_fast(enable_translations=True):
    """Fast CHF to EUR conversion for large files"""

    print("="*80)
    print("TXT File CHF to EUR Converter (FAST)")
    print("="*80)
    print()

    # Load databases once
    stock_df, omt_df, learning_db = load_databases()

    # Load input file
    print(f"[OK] Loaded input file: {INPUT_FILE_PATH}")
    with open(INPUT_FILE_PATH, 'r', encoding='utf-8') as f:
        text = f.read()

    # Find all CHF prices
    chf_matches = list(re.finditer(CHF_PRICE_PATTERN, text))
    print(f"[INFO] Found {len(chf_matches)} CHF prices in input file")
    print()

    # Process all prices
    recognized_items = []
    matched_omt_rows = []  # For OMT Excel export
    replacements = []  # (start, end, replacement_text)

    for i, match in enumerate(chf_matches, 1):
        # Extract price value
        price_str_raw = match.group(1) or match.group(2) or match.group(3)
        if not price_str_raw:
            continue

        # Remove thousands separator (')
        price_str = price_str_raw.replace("'", "")
        price_value = float(price_str)

        # Extract wine name and vintage from the line containing the price
        text_before = text[max(0, match.start()-200):match.start()]
        wine_name, vintage = extract_wine_name_vintage(text_before, text, match.start())

        # Try to match with min_quantity=0 first, then 36
        item_info = match_chf_to_eur(price_value, wine_name, vintage, stock_df, omt_df, learning_db, min_quantity=0)
        if item_info is None:
            item_info = match_chf_to_eur(price_value, wine_name, vintage, stock_df, omt_df, learning_db, min_quantity=36)

        if item_info:
            recognized_items.append(item_info)
            matched_omt_rows.append(item_info['omt_row'])

            eur_price = item_info['eur_price']
            eur_rounded = f"{int(round(float(eur_price)))}.00"

            # Create EUR replacement
            original_text = match.group(0)
            if 'CHF' in original_text:
                eur_text = original_text.replace('CHF', 'EUR')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            elif 'CH ' in original_text:
                eur_text = original_text.replace('CH ', 'EUR ')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            else:
                eur_text = original_text.replace(price_str_raw, eur_rounded)

            replacements.append((match.start(), match.end(), eur_text))

            safe_name = (wine_name or '').encode('ascii', errors='ignore').decode('ascii')
            print(f"[{i}/{len(chf_matches)}] Matched: {safe_name} {vintage} - CHF {price_value} -> EUR {eur_price} (Item {item_info['item_no']})")
        else:
            # Use fallback
            eur_fallback = price_value * 1.08
            eur_rounded = f"{int(math.floor(eur_fallback))}.00"

            original_text = match.group(0)
            if 'CHF' in original_text:
                eur_text = original_text.replace('CHF', 'EUR')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            else:
                eur_text = original_text.replace(price_str_raw, eur_rounded)

            replacements.append((match.start(), match.end(), eur_text))

            safe_name = (wine_name or '').encode('ascii', errors='ignore').decode('ascii')
            print(f"[{i}/{len(chf_matches)}] Fallback: {safe_name} {vintage} - CHF {price_value} -> EUR {eur_fallback:.2f} (no match)")

    # Apply all replacements in reverse order
    converted_text = text
    replacements.sort(key=lambda x: x[0], reverse=True)
    for start, end, replacement in replacements:
        converted_text = converted_text[:start] + replacement + converted_text[end:]

    # Generate timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Save converted text
    output_txt_path = rf"{OUTPUTS_DIR}\Multi_converted_{timestamp}.txt"
    with open(output_txt_path, 'w', encoding='utf-8') as f:
        f.write(converted_text)
    print(f"\n[OK] Output txt file saved: {output_txt_path}")

    # Save Stock Lines filtered Excel
    if recognized_items:
        unique_items = {}
        for item in recognized_items:
            item_no = item['item_no']
            if item_no not in unique_items:
                unique_items[item_no] = item['stock_row']

        filtered_df = pd.DataFrame([unique_items[k] for k in unique_items.keys()])
        stock_excel_path = rf"{RECOGNITION_REPORT_DIR}\Stock_Lines_Filtered_{timestamp}.xlsx"
        filtered_df.to_excel(stock_excel_path, index=False)
        print(f"[OK] Filtered Stock Lines Excel saved: {stock_excel_path}")
        print(f"   Contains {len(filtered_df)} recognized items")

    # Save Matched OMT Main Offer List Excel (Problem 3)
    if matched_omt_rows:
        Path(OMT_LINES_DIR).mkdir(parents=True, exist_ok=True)
        matched_omt_df = pd.DataFrame(matched_omt_rows)

        # Add extracted wine names from Multi.txt at the beginning
        extracted_names = []
        extracted_vintages = []
        for item in recognized_items:
            wine_name = item.get('wine_name', '')
            vintage = item.get('vintage', '')

            # Combine wine name and vintage
            if vintage and vintage != 'None':
                full_name = f"{wine_name} {vintage}"
            else:
                full_name = wine_name

            extracted_names.append(full_name)
            extracted_vintages.append(vintage if vintage != 'None' else '')

        # Insert new columns at the beginning
        matched_omt_df.insert(0, 'Extracted Vintage', extracted_vintages)
        matched_omt_df.insert(0, 'Extracted Wine Name (from Multi.txt)', extracted_names)

        matched_omt_path = rf"{OMT_LINES_DIR}\Matched_OMT Main Offer List_{timestamp}.xlsx"
        matched_omt_df.to_excel(matched_omt_path, index=False)
        print(f"[OK] Matched OMT Excel saved: {matched_omt_path}")
        print(f"   Contains {len(matched_omt_df)} matched OMT rows")

    # Generate Lines.xlsx for Business Central
    if recognized_items:
        try:
            Path(MAIN_OFFER_DIR).mkdir(parents=True, exist_ok=True)
            lines_excel_path = rf"{MAIN_OFFER_DIR}\Lines.xlsx"

            # Copy template
            copy2(LINES_TEMPLATE_PATH, lines_excel_path)

            # Fill data
            wb = openpyxl.load_workbook(lines_excel_path)
            ws = wb.active

            # Clear existing data rows (keep header)
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)

            # Write recognized items
            for idx, item in enumerate(recognized_items, start=2):
                stock_row = item['stock_row']
                omt_row = item['omt_row']

                ws.cell(idx, 1).value = stock_row.get('Wine Name')
                ws.cell(idx, 2).value = stock_row.get('Vintage Code')
                ws.cell(idx, 3).value = stock_row.get('Size')
                ws.cell(idx, 4).value = stock_row.get('Producer Name')
                ws.cell(idx, 5).value = omt_row.get('Minimum Quantity')
                ws.cell(idx, 6).value = item['chf_price']
                ws.cell(idx, 7).value = stock_row.get('Sales Price Base Inc. VAT (CHF)')
                ws.cell(idx, 8).value = int(round(float(item['eur_price'])))  # Rounded EUR
                ws.cell(idx, 9).value = omt_row.get('Main Offer Comment')
                ws.cell(idx, 10).value = None
                ws.cell(idx, 11).value = None

            wb.save(lines_excel_path)
            wb.close()

            print(f"[OK] Lines.xlsx saved with {len(recognized_items)} recognized wines")
            print(f"   Saved to: {lines_excel_path}")
        except Exception as e:
            print(f"[ERROR] Failed to generate Lines.xlsx: {e}")

    # Generate Recognition Report
    report_path = rf"{RECOGNITION_REPORT_DIR}\Recognition_Report_{timestamp}.txt"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("WINE RECOGNITION REPORT\n")
        f.write("="*80 + "\n\n")
        f.write(f"Total prices found: {len(chf_matches)}\n")
        f.write(f"Successfully matched: {len(recognized_items)}\n")
        f.write(f"Not matched: {len(chf_matches) - len(recognized_items)}\n\n")
        f.write("="*80 + "\n")
        f.write("RECOGNIZED WINES:\n")
        f.write("="*80 + "\n\n")

        for item in recognized_items:
            omt_row = item['omt_row']
            stock_row = item['stock_row']

            f.write(f"Wine (from Multi.txt): {item.get('wine_name') or 'N/A'}\n")
            f.write(f"Vintage: {item.get('vintage') or 'N/A'}\n")
            f.write(f"Database Match: {stock_row.get('Wine Name')}\n")
            f.write(f"CHF Price: {item['chf_price']}\n")
            f.write(f"EUR Price: {item['eur_price']}\n")
            f.write(f"Item No.: {item['item_no']}\n")
            f.write(f"Size: {omt_row.get('Size')}\n")
            f.write(f"Producer: {omt_row.get('Producer Name')}\n")
            f.write("-"*80 + "\n")

    print(f"[OK] Recognition report saved: {report_path}")

    # Statistics
    print("\n" + "="*80)
    print("CONVERSION STATISTICS")
    print("="*80)
    print(f"Total CHF prices found: {len(chf_matches)}")
    print(f"[OK] Successfully matched to Stock Lines: {len(recognized_items)}")
    print(f"[WARN] Not matched: {len(chf_matches) - len(recognized_items)}")
    print("="*80)
    print()

    # Translations
    if enable_translations:
        print("="*80)
        print("GENERATING TRANSLATIONS")
        print("="*80)
        print()

        Path(TRANSLATIONS_DIR).mkdir(parents=True, exist_ok=True)

        # German
        print("[INFO] Translating to German...")
        de_text = translate_text_deepl(converted_text, 'DE')
        if de_text:
            de_path = rf"{TRANSLATIONS_DIR}\Multi_DE_{timestamp}.txt"
            with open(de_path, 'w', encoding='utf-8') as f:
                f.write(de_text)
            print(f"[OK] German translation saved: {de_path}")

        # French
        print("[INFO] Translating to French...")
        fr_text = translate_text_deepl(converted_text, 'FR')
        if fr_text:
            fr_path = rf"{TRANSLATIONS_DIR}\Multi_FR_{timestamp}.txt"
            with open(fr_path, 'w', encoding='utf-8') as f:
                f.write(fr_text)
            print(f"[OK] French translation saved: {fr_path}")

        print("\n" + "="*80)
        print("CONVERSION AND TRANSLATION COMPLETE")
        print("="*80)
    else:
        print("\n[INFO] Translations skipped (checkbox disabled)")


if __name__ == "__main__":
    import sys
    enable_trans = True
    if len(sys.argv) > 1:
        enable_trans = sys.argv[1] == "1"
    convert_txt_file_fast(enable_translations=enable_trans)
