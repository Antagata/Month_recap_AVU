#!/usr/bin/env python3
"""
FAST Text File CHF to EUR Converter
Optimized for large files with 100+ CHF prices
"""

import re
import pandas as pd
from datetime import datetime
from pathlib import Path
import math
import openpyxl
from shutil import copy2
import requests

# Configuration
BASE_DIR = r"C:\Users\Marco.Africani\Desktop\Month recap"
DATABASE_DIR = r"C:\Users\Marco.Africani\OneDrive - AVU SA\AVU CPI Campaign\Puzzle_control_Reports\SOURCE_FILES"
INPUT_FILE_PATH = rf"{BASE_DIR}\Inputs\Multi.txt"
STOCK_FILE_PATH = rf"{DATABASE_DIR}\Stock Lines.xlsx"
OMT_FILE_PATH = rf"{DATABASE_DIR}\OMT Main Offer List.xlsx"
OUTPUTS_DIR = rf"{BASE_DIR}\Outputs"
RECOGNITION_REPORT_DIR = rf"{BASE_DIR}\Outputs\Detailed match results"
OMT_LINES_DIR = rf"{BASE_DIR}\Outputs\OMT lines"
TRANSLATIONS_DIR = rf"{BASE_DIR}\Outputs\EUR translations"
MAIN_OFFER_DIR = rf"{BASE_DIR}\Outputs\Detailed match results\Main offer"
LINES_TEMPLATE_PATH = rf"{MAIN_OFFER_DIR}\template\Lines Template.xlsx"

# DeepL API Configuration
DEEPL_API_KEY = "374a8965-101a-4538-bc65-54506552650e"
DEEPL_API_URL = "https://api.deepl.com/v2/translate"

# CHF price pattern - handles: "CHF 900.00", "230.00 CHF", "29.00 CH"
CHF_PRICE_PATTERN = r'(?:\bCHF\s+(\d+(?:[\']\d{3})*(?:\.\d{2})?)|(\d+(?:[\']\d{3})*(?:\.\d{2})?)\s+CHF\b|(\d+(?:[\']\d{3})*(?:\.\d{2})?)\s+CH\s)'


def load_databases():
    """Load Stock Lines and OMT databases once at startup"""
    print("Loading Stock Lines database...")
    stock_df = pd.read_excel(STOCK_FILE_PATH)
    print(f"[OK] Loaded {len(stock_df)} items from Stock Lines.xlsx")

    print("Loading OMT Main Offer List...")
    omt_df = pd.read_excel(OMT_FILE_PATH)
    print(f"[OK] Loaded {len(omt_df)} rows from OMT Main Offer List.xlsx")

    # Pre-process for faster lookups
    omt_df['Item No. Int'] = pd.to_numeric(omt_df['Item No.'], errors='coerce').astype('Int64')
    omt_df['Schedule DateTime'] = pd.to_datetime(omt_df['Schedule DateTime'])

    return stock_df, omt_df


def extract_wine_name_vintage(text_before_price):
    """Extract wine name and vintage from text before price"""
    # Look for "Wine Name VINTAGE:" or "Wine Name VINTAGE -"
    pattern = r'^\s*([^\d]+?)\s+(19\d{2}|20\d{2})\s*[^\w\s]'
    match = re.search(pattern, text_before_price[-150:])

    if match:
        wine_name = match.group(1).strip()
        vintage = int(match.group(2))
        # Clean up
        wine_name = re.sub(r'\s*\d+\s*$', '', wine_name)
        wine_name = re.sub(r'^(Château|Chateau|Domaine|Dom\.|Ch\.)\s+', '', wine_name, flags=re.IGNORECASE)
        return wine_name, vintage

    return None, None


def match_chf_to_eur(chf_price, wine_name, vintage, stock_df, omt_df, min_quantity=0):
    """
    BULLETPROOF: Stock Lines OMT Last Private Offer Price → OMT Unit Price + Min Quantity
    """
    try:
        chf_float = float(chf_price)

        # Step 1: Find in Stock Lines by OMT Last Private Offer Price
        stock_matches = stock_df[
            stock_df['OMT Last Private Offer Price'].astype(float).round(2) == round(chf_float, 2)
        ]

        if len(stock_matches) == 0:
            return None

        # Step 2: For each Stock Lines match, find in OMT
        all_candidates = []

        for _, stock_row in stock_matches.iterrows():
            item_no = stock_row['No.']

            # BULLETPROOF MATCH: Item No. + Unit Price (CHF) + Minimum Quantity
            omt_matches = omt_df[
                (omt_df['Item No. Int'] == item_no) &
                (omt_df['Unit Price'].astype(float).round(2) == round(chf_float, 2)) &
                (omt_df['Minimum Quantity'].astype(float) == float(min_quantity))
            ]

            for _, omt_row in omt_matches.iterrows():
                all_candidates.append({
                    'item_no': item_no,
                    'chf_price': chf_price,
                    'eur_price': omt_row.get('Unit Price (EUR)', 0),
                    'stock_row': stock_row,
                    'omt_row': omt_row,
                    'schedule_datetime': omt_row['Schedule DateTime'],
                    'wine_name': wine_name,
                    'vintage': vintage
                })

        if len(all_candidates) == 0:
            return None

        # Step 3: Pick latest Schedule DateTime
        all_candidates.sort(key=lambda x: x['schedule_datetime'], reverse=True)
        return all_candidates[0]

    except Exception as e:
        print(f"[WARN] Error matching CHF {chf_price}: {e}")
        return None


def translate_text_deepl(text, target_language):
    """Translate text using DeepL API"""
    try:
        params = {
            'auth_key': DEEPL_API_KEY,
            'text': text,
            'target_lang': target_language
        }
        response = requests.post(DEEPL_API_URL, data=params, timeout=30)

        if response.status_code == 200:
            return response.json()['translations'][0]['text']
        else:
            print(f"[WARN] DeepL API error {response.status_code}: {response.text}")
            return None
    except Exception as e:
        print(f"[WARN] Translation error: {e}")
        return None


def convert_txt_file_fast(enable_translations=True):
    """Fast CHF to EUR conversion for large files"""

    print("="*80)
    print("TXT File CHF to EUR Converter (FAST)")
    print("="*80)
    print()

    # Load databases once
    stock_df, omt_df = load_databases()

    # Load input file
    print(f"[OK] Loaded input file: {INPUT_FILE_PATH}")
    with open(INPUT_FILE_PATH, 'r', encoding='utf-8') as f:
        text = f.read()

    # Find all CHF prices
    chf_matches = list(re.finditer(CHF_PRICE_PATTERN, text))
    print(f"[INFO] Found {len(chf_matches)} CHF prices in input file")
    print()

    # Process all prices
    recognized_items = []
    matched_omt_rows = []  # For OMT Excel export
    replacements = []  # (start, end, replacement_text)

    for i, match in enumerate(chf_matches, 1):
        # Extract price value
        price_str_raw = match.group(1) or match.group(2) or match.group(3)
        if not price_str_raw:
            continue

        # Remove thousands separator (')
        price_str = price_str_raw.replace("'", "")
        price_value = float(price_str)

        # Get text before price for wine name extraction
        text_before = text[max(0, match.start()-200):match.start()]
        wine_name, vintage = extract_wine_name_vintage(text_before)

        # Try to match with min_quantity=0 first, then 36
        item_info = match_chf_to_eur(price_value, wine_name, vintage, stock_df, omt_df, min_quantity=0)
        if item_info is None:
            item_info = match_chf_to_eur(price_value, wine_name, vintage, stock_df, omt_df, min_quantity=36)

        if item_info:
            recognized_items.append(item_info)
            matched_omt_rows.append(item_info['omt_row'])

            eur_price = item_info['eur_price']
            eur_rounded = f"{int(round(float(eur_price)))}.00"

            # Create EUR replacement
            original_text = match.group(0)
            if 'CHF' in original_text:
                eur_text = original_text.replace('CHF', 'EUR')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            elif 'CH ' in original_text:
                eur_text = original_text.replace('CH ', 'EUR ')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            else:
                eur_text = original_text.replace(price_str_raw, eur_rounded)

            replacements.append((match.start(), match.end(), eur_text))

            safe_name = (wine_name or '').encode('ascii', errors='ignore').decode('ascii')
            print(f"[{i}/{len(chf_matches)}] Matched: {safe_name} {vintage} - CHF {price_value} -> EUR {eur_price} (Item {item_info['item_no']})")
        else:
            # Use fallback
            eur_fallback = price_value * 1.08
            eur_rounded = f"{int(math.floor(eur_fallback))}.00"

            original_text = match.group(0)
            if 'CHF' in original_text:
                eur_text = original_text.replace('CHF', 'EUR')
                eur_text = re.sub(r"\d+(?:[\']\d{3})*(?:\.\d{2})?", eur_rounded, eur_text)
            else:
                eur_text = original_text.replace(price_str_raw, eur_rounded)

            replacements.append((match.start(), match.end(), eur_text))

            safe_name = (wine_name or '').encode('ascii', errors='ignore').decode('ascii')
            print(f"[{i}/{len(chf_matches)}] Fallback: {safe_name} {vintage} - CHF {price_value} -> EUR {eur_fallback:.2f} (no match)")

    # Apply all replacements in reverse order
    converted_text = text
    replacements.sort(key=lambda x: x[0], reverse=True)
    for start, end, replacement in replacements:
        converted_text = converted_text[:start] + replacement + converted_text[end:]

    # Generate timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Save converted text
    output_txt_path = rf"{OUTPUTS_DIR}\Multi_converted_{timestamp}.txt"
    with open(output_txt_path, 'w', encoding='utf-8') as f:
        f.write(converted_text)
    print(f"\n[OK] Output txt file saved: {output_txt_path}")

    # Save Stock Lines filtered Excel
    if recognized_items:
        unique_items = {}
        for item in recognized_items:
            item_no = item['item_no']
            if item_no not in unique_items:
                unique_items[item_no] = item['stock_row']

        filtered_df = pd.DataFrame([unique_items[k] for k in unique_items.keys()])
        stock_excel_path = rf"{RECOGNITION_REPORT_DIR}\Stock_Lines_Filtered_{timestamp}.xlsx"
        filtered_df.to_excel(stock_excel_path, index=False)
        print(f"[OK] Filtered Stock Lines Excel saved: {stock_excel_path}")
        print(f"   Contains {len(filtered_df)} recognized items")

    # Save Matched OMT Main Offer List Excel (Problem 3)
    if matched_omt_rows:
        Path(OMT_LINES_DIR).mkdir(parents=True, exist_ok=True)
        matched_omt_df = pd.DataFrame(matched_omt_rows)
        matched_omt_path = rf"{OMT_LINES_DIR}\Matched_OMT Main Offer List_{timestamp}.xlsx"
        matched_omt_df.to_excel(matched_omt_path, index=False)
        print(f"[OK] Matched OMT Excel saved: {matched_omt_path}")
        print(f"   Contains {len(matched_omt_df)} matched OMT rows")

    # Generate Lines.xlsx for Business Central
    if recognized_items:
        try:
            Path(MAIN_OFFER_DIR).mkdir(parents=True, exist_ok=True)
            lines_excel_path = rf"{MAIN_OFFER_DIR}\Lines.xlsx"

            # Copy template
            copy2(LINES_TEMPLATE_PATH, lines_excel_path)

            # Fill data
            wb = openpyxl.load_workbook(lines_excel_path)
            ws = wb.active

            # Clear existing data rows (keep header)
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)

            # Write recognized items
            for idx, item in enumerate(recognized_items, start=2):
                stock_row = item['stock_row']
                omt_row = item['omt_row']

                ws.cell(idx, 1).value = stock_row.get('Wine Name')
                ws.cell(idx, 2).value = stock_row.get('Vintage Code')
                ws.cell(idx, 3).value = stock_row.get('Size')
                ws.cell(idx, 4).value = stock_row.get('Producer Name')
                ws.cell(idx, 5).value = omt_row.get('Minimum Quantity')
                ws.cell(idx, 6).value = item['chf_price']
                ws.cell(idx, 7).value = stock_row.get('Sales Price Base Inc. VAT (CHF)')
                ws.cell(idx, 8).value = int(round(float(item['eur_price'])))  # Rounded EUR
                ws.cell(idx, 9).value = omt_row.get('Main Offer Comment')
                ws.cell(idx, 10).value = None
                ws.cell(idx, 11).value = None

            wb.save(lines_excel_path)
            wb.close()

            print(f"[OK] Lines.xlsx saved with {len(recognized_items)} recognized wines")
            print(f"   Saved to: {lines_excel_path}")
        except Exception as e:
            print(f"[ERROR] Failed to generate Lines.xlsx: {e}")

    # Generate Recognition Report
    report_path = rf"{RECOGNITION_REPORT_DIR}\Recognition_Report_{timestamp}.txt"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("WINE RECOGNITION REPORT\n")
        f.write("="*80 + "\n\n")
        f.write(f"Total prices found: {len(chf_matches)}\n")
        f.write(f"Successfully matched: {len(recognized_items)}\n")
        f.write(f"Not matched: {len(chf_matches) - len(recognized_items)}\n\n")
        f.write("="*80 + "\n")
        f.write("RECOGNIZED WINES:\n")
        f.write("="*80 + "\n\n")

        for item in recognized_items:
            omt_row = item['omt_row']
            stock_row = item['stock_row']

            f.write(f"Wine (from Multi.txt): {item.get('wine_name') or 'N/A'}\n")
            f.write(f"Vintage: {item.get('vintage') or 'N/A'}\n")
            f.write(f"Database Match: {stock_row.get('Wine Name')}\n")
            f.write(f"CHF Price: {item['chf_price']}\n")
            f.write(f"EUR Price: {item['eur_price']}\n")
            f.write(f"Item No.: {item['item_no']}\n")
            f.write(f"Size: {omt_row.get('Size')}\n")
            f.write(f"Producer: {omt_row.get('Producer Name')}\n")
            f.write("-"*80 + "\n")

    print(f"[OK] Recognition report saved: {report_path}")

    # Statistics
    print("\n" + "="*80)
    print("CONVERSION STATISTICS")
    print("="*80)
    print(f"Total CHF prices found: {len(chf_matches)}")
    print(f"[OK] Successfully matched to Stock Lines: {len(recognized_items)}")
    print(f"[WARN] Not matched: {len(chf_matches) - len(recognized_items)}")
    print("="*80)
    print()

    # Translations
    if enable_translations:
        print("="*80)
        print("GENERATING TRANSLATIONS")
        print("="*80)
        print()

        Path(TRANSLATIONS_DIR).mkdir(parents=True, exist_ok=True)

        # German
        print("[INFO] Translating to German...")
        de_text = translate_text_deepl(converted_text, 'DE')
        if de_text:
            de_path = rf"{TRANSLATIONS_DIR}\Multi_DE_{timestamp}.txt"
            with open(de_path, 'w', encoding='utf-8') as f:
                f.write(de_text)
            print(f"[OK] German translation saved: {de_path}")

        # French
        print("[INFO] Translating to French...")
        fr_text = translate_text_deepl(converted_text, 'FR')
        if fr_text:
            fr_path = rf"{TRANSLATIONS_DIR}\Multi_FR_{timestamp}.txt"
            with open(fr_path, 'w', encoding='utf-8') as f:
                f.write(fr_text)
            print(f"[OK] French translation saved: {fr_path}")

        print("\n" + "="*80)
        print("CONVERSION AND TRANSLATION COMPLETE")
        print("="*80)
    else:
        print("\n[INFO] Translations skipped (checkbox disabled)")


if __name__ == "__main__":
    import sys
    enable_trans = True
    if len(sys.argv) > 1:
        enable_trans = sys.argv[1] == "1"
    convert_txt_file_fast(enable_translations=enable_trans)
